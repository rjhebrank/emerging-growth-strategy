"""
Build Bloomberg Excel Template for Emerging Growth Strategy.

Generates a 6-sheet workbook with Bloomberg formulas (BDS/BDP/BDH)
saved as plain text strings. Bloomberg evaluates them when opened on
a terminal machine with the Excel Add-in active.

Usage:
    source .venv/bin/activate && python scripts/build_template.py
"""

import datetime
from copy import copy

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
FONT_DEFAULT = Font(name="Calibri", size=11)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FONT_GREEN = Font(name="Calibri", size=11, color="006100")
FONT_RED = Font(name="Calibri", size=11, color="9C0006")
FONT_BLUE = Font(name="Calibri", size=11, color="0000FF")

TAB_COLORS = {
    "README": "808080",
    "Universe": "4472C4",
    "Price History": "70AD47",
    "Fundamentals": "ED7D31",
    "Factor Scores": "7030A0",
    "Portfolio": "FF0000",
}

OUTPUT_PATH = "/home/riley/emerging-growth-strategy/templates/bloomberg_template.xlsx"

MAX_DATA_ROW = 2501  # rows 2-2501 = 2500 data rows


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def apply_header_style(ws, row, col_count):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def set_col_widths(ws, widths: dict):
    """Set column widths from {column_letter: width} dict."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_row(ws, row, values, font=None, bold=False):
    """Write a list of values to a row starting at column A."""
    f = font or Font(name="Calibri", size=11, bold=bold)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = f


# ---------------------------------------------------------------------------
# Sheet 0: README
# ---------------------------------------------------------------------------
def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = TAB_COLORS["README"]

    # Merge title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Emerging Growth Strategy — Bloomberg Data Template"
    title_cell.font = Font(name="Calibri", size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    today_str = datetime.date.today().strftime("%B %d, %Y")
    ws["A3"] = f"Generated: {today_str}"
    ws["A3"].font = Font(name="Calibri", size=11)

    # FIRST-TIME SETUP
    ws["A5"] = "FIRST-TIME SETUP"
    ws["A5"].font = Font(name="Calibri", size=11, bold=True)

    setup_steps = [
        '1. Open this file on a computer with Bloomberg Terminal and Excel Add-in active',
        '2. Go to the Universe sheet — the BDS formula in A2 will pull ~2,000 Russell 2000 tickers',
        '3. Press Ctrl+Shift+R to refresh all Bloomberg formulas (wait 5-10 minutes)',
        '4. After Universe populates, filter column K for TRUE to find qualifying tickers',
        '5. Copy qualifying tickers to the "Ticker Input" columns on Price History (col I) and Fundamentals (col M) sheets',
        '6. Press Ctrl+Shift+R again to populate price and fundamental data',
        '7. Factor Scores and Portfolio sheets auto-calculate from the data',
    ]
    for i, step in enumerate(setup_steps):
        ws.cell(row=6 + i, column=1, value=step).font = FONT_DEFAULT

    # MONTHLY REFRESH
    ws["A14"] = "MONTHLY REFRESH"
    ws["A14"].font = Font(name="Calibri", size=11, bold=True)

    refresh_steps = [
        '1. Open workbook on first trading day of month',
        '2. Dates in BDH formulas use TODAY() — they auto-update',
        '3. Press Ctrl+Shift+R to refresh all Bloomberg data',
        '4. Check for #N/A errors — use alternative field codes from docs/06-bloomberg-data-pull.md',
        '5. Save a dated copy (File → Save As → "Strategy_YYYY-MM-DD.xlsx")',
    ]
    for i, step in enumerate(refresh_steps):
        ws.cell(row=15 + i, column=1, value=step).font = FONT_DEFAULT

    # TROUBLESHOOTING
    ws["A21"] = "TROUBLESHOOTING"
    ws["A21"].font = Font(name="Calibri", size=11, bold=True)

    troubleshooting = [
        '#NAME? errors: Bloomberg Excel Add-in not loaded. Restart Excel with Bloomberg Terminal open.',
        '#N/A errors: Field unavailable for that ticker. Try alternative fields listed in docs.',
        'Slow refresh: Pull in batches — too many simultaneous Bloomberg requests cause timeouts.',
        'See docs/06-bloomberg-data-pull.md for complete reference.',
    ]
    for i, line in enumerate(troubleshooting):
        ws.cell(row=22 + i, column=1, value=line).font = FONT_DEFAULT

    # Widen column A so text is readable
    ws.column_dimensions["A"].width = 100


# ---------------------------------------------------------------------------
# Sheet 1: Universe
# ---------------------------------------------------------------------------
def build_universe(wb):
    ws = wb.create_sheet("Universe")
    ws.sheet_properties.tabColor = TAB_COLORS["Universe"]

    headers = [
        "Ticker", "Market Cap ($M)", "Exchange", "Avg Vol (20d)",
        "Price", "Dollar Volume", "MktCap Pass", "Exchange Pass",
        "DolVol Pass", "Price Pass", "ALL PASS",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws, 1, len(headers))

    # Row 2 — seed formulas
    ws["A2"] = '=BDS("RTY Index","INDX_MEMBERS")'
    ws["B2"] = '=BDP(A2&" US Equity","CUR_MKT_CAP")'
    ws["C2"] = '=BDP(A2&" US Equity","EXCH_CODE")'
    ws["D2"] = '=BDP(A2&" US Equity","VOLUME_AVG_20D")'
    ws["E2"] = '=BDP(A2&" US Equity","PX_LAST")'
    ws["F2"] = "=D2*E2"
    ws["G2"] = "=AND(B2>=50,B2<=10000)"
    ws["H2"] = '=OR(C2="NAS",C2="NYS",C2="ASE")'
    ws["I2"] = "=F2>=500000"
    ws["J2"] = "=E2>=2"
    ws["K2"] = "=AND(G2,H2,I2,J2)"

    # Rows 3-2501: B-K formulas only (A column left blank for BDS spill)
    for r in range(3, MAX_DATA_ROW + 1):
        ws[f"B{r}"] = f'=BDP(A{r}&" US Equity","CUR_MKT_CAP")'
        ws[f"C{r}"] = f'=BDP(A{r}&" US Equity","EXCH_CODE")'
        ws[f"D{r}"] = f'=BDP(A{r}&" US Equity","VOLUME_AVG_20D")'
        ws[f"E{r}"] = f'=BDP(A{r}&" US Equity","PX_LAST")'
        ws[f"F{r}"] = f"=D{r}*E{r}"
        ws[f"G{r}"] = f"=AND(B{r}>=50,B{r}<=10000)"
        ws[f"H{r}"] = f'=OR(C{r}="NAS",C{r}="NYS",C{r}="ASE")'
        ws[f"I{r}"] = f"=F{r}>=500000"
        ws[f"J{r}"] = f"=E{r}>=2"
        ws[f"K{r}"] = f"=AND(G{r},H{r},I{r},J{r})"

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 16, "C": 10, "D": 16, "E": 10, "F": 16,
        "G": 13, "H": 13, "I": 13, "J": 13, "K": 13,
    })

    # Number formats
    for r in range(2, MAX_DATA_ROW + 1):
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"D{r}"].number_format = "#,##0"
        ws[f"E{r}"].number_format = "#,##0.00"
        ws[f"F{r}"].number_format = "#,##0"

    # Conditional formatting on column K
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        f"K2:K{MAX_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"TRUE"'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"K2:K{MAX_DATA_ROW}",
        FormulaRule(formula=["K2=TRUE"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"K2:K{MAX_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"FALSE"'], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"K2:K{MAX_DATA_ROW}",
        FormulaRule(formula=["K2=FALSE"], fill=red_fill),
    )

    # Freeze pane
    ws.freeze_panes = "A2"

    print("  Universe sheet: done")


# ---------------------------------------------------------------------------
# Sheet 2: Price History
# ---------------------------------------------------------------------------
def build_price_history(wb):
    ws = wb.create_sheet("Price History")
    ws.sheet_properties.tabColor = TAB_COLORS["Price History"]

    # Instruction row
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "INSTRUCTIONS: Paste qualifying tickers into column I (Ticker Input). "
        "BDH formulas auto-populate OHLCV data. Each ticker block = 320 rows."
    )
    ws["A1"].font = Font(name="Calibri", size=11, bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True)

    # Row 2 headers
    headers = {
        "A": "Ticker", "B": "Date", "C": "Open", "D": "High",
        "E": "Low", "F": "Close", "G": "Volume", "I": "Ticker Input",
    }
    for col_letter, h in headers.items():
        ws[f"{col_letter}2"] = h
    # Apply header style to row 2 (A-G and I)
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "I"]:
        cell = ws[f"{col_letter}2"]
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # BDH formula template
    bdh_template = (
        '=BDH(A{row}&" US Equity",'
        '"PX_OPEN,PX_HIGH,PX_LOW,PX_LAST,VOLUME",'
        'TEXT(EDATE(TODAY(),-15),"MM/DD/YYYY"),'
        'TEXT(TODAY(),"MM/DD/YYYY"),'
        '"Days","A","Fill","P","CshAdjNormal","Y","CshAdjAbnormal","Y")'
    )

    # 100 ticker blocks, each 320 rows
    for i in range(100):
        start_row = 3 + i * 320
        ticker_input_row = 3 + i  # I3, I4, I5, ...

        ws[f"A{start_row}"] = f"=I{ticker_input_row}"
        ws[f"B{start_row}"] = bdh_template.format(row=start_row)

    # Yellow fill on column I (ticker input area) rows 3-102
    for r in range(3, 103):
        ws[f"I{r}"].fill = FILL_YELLOW

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 12, "C": 10, "D": 10, "E": 10, "F": 10, "G": 12, "I": 14,
    })

    # Number formats for C-F and G across all potential data rows
    last_data_row = 3 + 100 * 320  # 32003
    for i in range(100):
        start_row = 3 + i * 320
        # Apply number format to the first row of each block (BDH spills the rest)
        for col in ["C", "D", "E", "F"]:
            ws[f"{col}{start_row}"].number_format = "#,##0.00"
        ws[f"G{start_row}"].number_format = "#,##0"

    # Freeze pane at A3
    ws.freeze_panes = "A3"

    print("  Price History sheet: done")


# ---------------------------------------------------------------------------
# Sheet 3: Fundamentals
# ---------------------------------------------------------------------------
def build_fundamentals(wb):
    ws = wb.create_sheet("Fundamentals")
    ws.sheet_properties.tabColor = TAB_COLORS["Fundamentals"]

    # Row 1 headers
    header_map = {
        "A": "Ticker", "B": "EPS Date", "C": "EPS",
        "J": "Rev Date", "K": "Revenue", "L": "EPS YoY %",
        "M": "Ticker Input", "N": "Rev YoY %",
    }
    for col_letter, h in header_map.items():
        cell = ws[f"{col_letter}1"]
        cell.value = h
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Also style D-I headers (empty but should have header fill for continuity)
    for col_letter in ["D", "E", "F", "G", "H", "I"]:
        cell = ws[f"{col_letter}1"]
        cell.fill = FILL_HEADER

    # BDH templates
    eps_bdh = (
        '=BDH(A{row}&" US Equity","IS_DILUTED_EPS",'
        'TEXT(EDATE(TODAY(),-24),"MM/DD/YYYY"),'
        'TEXT(TODAY(),"MM/DD/YYYY"),'
        '"Per","Q","Days","A","Fill","P")'
    )
    rev_bdh = (
        '=BDH(A{row}&" US Equity","SALES_REV_TURN",'
        'TEXT(EDATE(TODAY(),-24),"MM/DD/YYYY"),'
        'TEXT(TODAY(),"MM/DD/YYYY"),'
        '"Per","Q","Days","A","Fill","P")'
    )

    # 100 ticker blocks, each 10 rows starting at row 2
    for i in range(100):
        sr = 2 + i * 10  # start row
        ticker_input_row = 2 + i  # M2, M3, M4, ...

        ws[f"A{sr}"] = f"=M{ticker_input_row}"
        ws[f"B{sr}"] = eps_bdh.format(row=sr)
        ws[f"J{sr}"] = rev_bdh.format(row=sr)

        # EPS YoY: compare row sr+7 (newest) vs sr+3 (4 quarters ago)
        # Guard division by zero and negative denominators
        eps_newest = f"C{sr+7}"
        eps_old = f"C{sr+3}"
        ws[f"L{sr}"] = (
            f"=IF({eps_old}=0,IF({eps_newest}>0,999,0),"
            f"IF({eps_old}<0,IF({eps_newest}>0,999,0),"
            f"({eps_newest}-{eps_old})/ABS({eps_old})*100))"
        )

        # Rev YoY
        rev_newest = f"K{sr+7}"
        rev_old = f"K{sr+3}"
        ws[f"N{sr}"] = (
            f"=IF({rev_old}<=0,IF({rev_newest}>0,999,0),"
            f"({rev_newest}-{rev_old})/{rev_old}*100)"
        )

    # Yellow fill on column M (ticker input area) rows 2-101
    for r in range(2, 102):
        ws[f"M{r}"].fill = FILL_YELLOW

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 12, "C": 10, "J": 12, "K": 14, "L": 12, "M": 14, "N": 12,
    })

    # Number formats
    for i in range(100):
        sr = 2 + i * 10
        ws[f"C{sr}"].number_format = "#,##0.00"
        ws[f"K{sr}"].number_format = "#,##0"
        ws[f"L{sr}"].number_format = "0.00"
        ws[f"N{sr}"].number_format = "0.00"

    # Freeze pane — freeze below row 1 headers
    ws.freeze_panes = "A2"

    print("  Fundamentals sheet: done")


# ---------------------------------------------------------------------------
# Sheet 4: Factor Scores
# ---------------------------------------------------------------------------
def build_factor_scores(wb):
    ws = wb.create_sheet("Factor Scores")
    ws.sheet_properties.tabColor = TAB_COLORS["Factor Scores"]

    headers = [
        "Ticker", "6-Mo Return (%)", "RS Percentile",
        "EPS Growth YoY", "EPS (capped)", "Rev Growth YoY", "Rev (capped)",
        "Price vs 52Wk High (%)", "Qual: EPS≥5%", "Qual: Rev≥5%",
        "Qual: Price≥75%", "ALL QUALITY PASS",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws, 1, len(headers))

    for r in range(2, MAX_DATA_ROW + 1):
        ws[f"A{r}"] = f"=Universe!A{r}"

        # B: placeholder — requires complex cross-sheet lookup from Price History
        # Write a cell comment instead
        ws[f"B{r}"].value = None

        ws[f"C{r}"] = f'=IF(B{r}="","",PERCENTRANK.INC($B$2:$B$2501,B{r})*100)'

        # D: placeholder for EPS growth linkage
        ws[f"D{r}"].value = None

        ws[f"E{r}"] = f'=IF(D{r}="","",MIN(D{r},100))'

        # F: placeholder for Revenue growth linkage
        ws[f"F{r}"].value = None

        ws[f"G{r}"] = f'=IF(F{r}="","",MIN(F{r},100))'

        ws[f"H{r}"] = (
            f'=IF(A{r}="","",Universe!E{r}/BDP(A{r}&" US Equity","HIGH_52WEEK")*100)'
        )

        ws[f"I{r}"] = f'=IF(D{r}="","",D{r}>=5)'
        ws[f"J{r}"] = f'=IF(F{r}="","",F{r}>=5)'
        ws[f"K{r}"] = f'=IF(H{r}="","",H{r}>=75)'
        ws[f"L{r}"] = f'=IF(OR(I{r}="",J{r}="",K{r}=""),FALSE,AND(I{r},J{r},K{r}))'

    # Add notes to B2 and D2, F2 explaining the placeholders
    from openpyxl.comments import Comment
    ws["B2"].comment = Comment(
        "Populate from Python engine or Price History sheet. "
        "6-month return for each ticker.",
        "Template Builder",
    )
    ws["D2"].comment = Comment(
        "Link to Fundamentals sheet L column for corresponding ticker.",
        "Template Builder",
    )
    ws["F2"].comment = Comment(
        "Link to Fundamentals sheet N column for corresponding ticker.",
        "Template Builder",
    )

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 14, "C": 13, "D": 14, "E": 12, "F": 14,
        "G": 12, "H": 20, "I": 13, "J": 13, "K": 14, "L": 17,
    })

    # Number formats
    for r in range(2, MAX_DATA_ROW + 1):
        ws[f"B{r}"].number_format = "0.00"
        ws[f"C{r}"].number_format = "0.0"
        ws[f"D{r}"].number_format = "0.00"
        ws[f"E{r}"].number_format = "0.00"
        ws[f"F{r}"].number_format = "0.00"
        ws[f"G{r}"].number_format = "0.00"
        ws[f"H{r}"].number_format = "0.0"

    # Conditional formatting on L column
    ws.conditional_formatting.add(
        f"L2:L{MAX_DATA_ROW}",
        FormulaRule(formula=["L2=TRUE"], fill=FILL_GREEN),
    )
    ws.conditional_formatting.add(
        f"L2:L{MAX_DATA_ROW}",
        FormulaRule(formula=["L2=FALSE"], fill=FILL_RED),
    )

    # Freeze pane
    ws.freeze_panes = "A2"

    print("  Factor Scores sheet: done")


# ---------------------------------------------------------------------------
# Sheet 5: Portfolio
# ---------------------------------------------------------------------------
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio")
    ws.sheet_properties.tabColor = TAB_COLORS["Portfolio"]

    headers = [
        "Ticker", "Composite Score", "Rank", "Top 25?",
        "Prior Month", "Signal", "Weight (%)", "Sector",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    apply_header_style(ws, 1, len(headers))

    for r in range(2, MAX_DATA_ROW + 1):
        ws[f"A{r}"] = f"='Factor Scores'!A{r}"
        ws[f"B{r}"] = (
            f"=IF(A{r}=\"\",\"\","
            f"0.40*'Factor Scores'!C{r}+"
            f"0.20*'Factor Scores'!E{r}+"
            f"0.20*'Factor Scores'!G{r}+"
            f"0.20*'Factor Scores'!H{r})"
        )
        ws[f"C{r}"] = f'=IF(B{r}="","",RANK(B{r},$B$2:$B$2501,0))'
        ws[f"D{r}"] = f'=IF(C{r}="","",C{r}<=25)'
        # E: Prior Month — left blank for user
        ws[f"F{r}"] = (
            f'=IF(A{r}="","",IF(AND(D{r}=TRUE,E{r}=""),"BUY",'
            f'IF(AND(D{r}=FALSE,E{r}<>""),"SELL",'
            f'IF(AND(D{r}=TRUE,E{r}<>""),"HOLD",""))))'
        )
        ws[f"G{r}"] = f"=IF(D{r}=TRUE,4,0)"
        ws[f"H{r}"] = f'=IF(A{r}="","",BDP(A{r}&" US Equity","GICS_SECTOR_NAME"))'

    # Column widths
    set_col_widths(ws, {
        "A": 12, "B": 16, "C": 8, "D": 10, "E": 14, "F": 10, "G": 12, "H": 20,
    })

    # Number formats
    for r in range(2, MAX_DATA_ROW + 1):
        ws[f"B{r}"].number_format = "0.00"
        ws[f"C{r}"].number_format = "0"
        ws[f"G{r}"].number_format = "0"

    # Conditional formatting on D column (TRUE/FALSE)
    ws.conditional_formatting.add(
        f"D2:D{MAX_DATA_ROW}",
        FormulaRule(formula=["D2=TRUE"], fill=FILL_GREEN),
    )
    ws.conditional_formatting.add(
        f"D2:D{MAX_DATA_ROW}",
        FormulaRule(formula=["D2=FALSE"], fill=FILL_RED),
    )

    # Conditional formatting on F column (BUY/SELL/HOLD text color)
    ws.conditional_formatting.add(
        f"F2:F{MAX_DATA_ROW}",
        CellIsRule(
            operator="equal",
            formula=['"BUY"'],
            font=Font(name="Calibri", size=11, bold=True, color="006100"),
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        f"F2:F{MAX_DATA_ROW}",
        CellIsRule(
            operator="equal",
            formula=['"SELL"'],
            font=Font(name="Calibri", size=11, bold=True, color="9C0006"),
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        f"F2:F{MAX_DATA_ROW}",
        CellIsRule(
            operator="equal",
            formula=['"HOLD"'],
            font=Font(name="Calibri", size=11, bold=True, color="0000FF"),
            fill=PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        ),
    )

    # Freeze pane
    ws.freeze_panes = "A2"

    print("  Portfolio sheet: done")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Building Bloomberg template...")
    wb = Workbook()

    build_readme(wb)
    print("  README sheet: done")

    build_universe(wb)
    build_price_history(wb)
    build_fundamentals(wb)
    build_factor_scores(wb)
    build_portfolio(wb)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)

    import os
    size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
    print(f"Done! File size: {size_mb:.1f} MB")


if __name__ == "__main__":
    main()

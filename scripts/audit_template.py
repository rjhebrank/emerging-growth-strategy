"""
Audit Bloomberg Excel Template.

Opens the generated template and validates every formula on every sheet
against the specification. Reports all findings.

Usage:
    source .venv/bin/activate && python scripts/audit_template.py
"""

import random
import sys

from openpyxl import load_workbook

TEMPLATE_PATH = "/home/riley/emerging-growth-strategy/templates/bloomberg_template.xlsx"

# Track results
errors = []
warnings = []
passes = []


def log_pass(msg):
    passes.append(msg)
    print(f"  [PASS] {msg}")


def log_error(msg):
    errors.append(msg)
    print(f"  [ERROR] {msg}")


def log_warn(msg):
    warnings.append(msg)
    print(f"  [WARN] {msg}")


def get_formula(ws, cell_ref):
    """Return the value of a cell as a string. Formulas start with '='."""
    val = ws[cell_ref].value
    return str(val) if val is not None else ""


# ---------------------------------------------------------------------------
# Universe Sheet Audit
# ---------------------------------------------------------------------------
def audit_universe(wb):
    print("\n=== Auditing Universe Sheet ===")
    ws = wb["Universe"]

    # A2 check
    a2 = get_formula(ws, "A2")
    if a2 == '=BDS("RTY Index","INDX_MEMBERS")':
        log_pass("A2: BDS formula correct")
    else:
        log_error(f'A2: Expected =BDS("RTY Index","INDX_MEMBERS"), got: {a2}')

    # B2 check
    b2 = get_formula(ws, "B2")
    if "BDP" in b2 and "CUR_MKT_CAP" in b2:
        log_pass("B2: BDP with CUR_MKT_CAP correct")
    else:
        log_error(f"B2: Expected BDP with CUR_MKT_CAP, got: {b2}")

    # C2 check
    c2 = get_formula(ws, "C2")
    if "BDP" in c2 and "EXCH_CODE" in c2:
        log_pass("C2: BDP with EXCH_CODE correct")
    else:
        log_error(f"C2: Expected BDP with EXCH_CODE, got: {c2}")

    # D2 check
    d2 = get_formula(ws, "D2")
    if "BDP" in d2 and "VOLUME_AVG_20D" in d2:
        log_pass("D2: BDP with VOLUME_AVG_20D correct")
    else:
        log_error(f"D2: Expected BDP with VOLUME_AVG_20D, got: {d2}")

    # E2 check
    e2 = get_formula(ws, "E2")
    if "BDP" in e2 and "PX_LAST" in e2:
        log_pass("E2: BDP with PX_LAST correct")
    else:
        log_error(f"E2: Expected BDP with PX_LAST, got: {e2}")

    # F2 check
    f2 = get_formula(ws, "F2")
    if f2 == "=D2*E2":
        log_pass("F2: Dollar volume formula =D2*E2 correct")
    else:
        log_error(f"F2: Expected =D2*E2, got: {f2}")

    # G2 check
    g2 = get_formula(ws, "G2")
    if "AND" in g2 and "B2>=50" in g2 and "B2<=10000" in g2:
        log_pass("G2: Market cap filter 50-10000 correct")
    else:
        log_error(f"G2: Expected AND(B2>=50,B2<=10000), got: {g2}")

    # H2 check
    h2 = get_formula(ws, "H2")
    if "OR" in h2 and "NAS" in h2 and "NYS" in h2 and "ASE" in h2:
        log_pass("H2: Exchange filter correct")
    else:
        log_error(f"H2: Expected OR(C2=NAS/NYS/ASE), got: {h2}")

    # I2 check
    i2 = get_formula(ws, "I2")
    if "F2>=500000" in i2:
        log_pass("I2: Dollar volume >= 500000 check correct")
    else:
        log_error(f"I2: Expected =F2>=500000, got: {i2}")

    # J2 check
    j2 = get_formula(ws, "J2")
    if "E2>=2" in j2:
        log_pass("J2: Price >= 2 check correct")
    else:
        log_error(f"J2: Expected =E2>=2, got: {j2}")

    # K2 check
    k2 = get_formula(ws, "K2")
    if "AND" in k2 and "G2" in k2 and "H2" in k2 and "I2" in k2 and "J2" in k2:
        log_pass("K2: AND(G2,H2,I2,J2) correct")
    else:
        log_error(f"K2: Expected AND(G2,H2,I2,J2), got: {k2}")

    # CRITICAL: Check random rows to verify formulas reference their own row
    random.seed(42)
    sample_rows = random.sample(range(3, 2502), 10)
    print(f"\n  Checking row-reference integrity for rows: {sorted(sample_rows)}")

    all_rows_ok = True
    for r in sorted(sample_rows):
        row_errors = []

        # B should reference A{r}
        b_val = get_formula(ws, f"B{r}")
        if f"A{r}" not in b_val:
            row_errors.append(f"B{r} does not reference A{r}: {b_val}")

        # F should be D{r}*E{r}
        f_val = get_formula(ws, f"F{r}")
        if f"D{r}" not in f_val or f"E{r}" not in f_val:
            row_errors.append(f"F{r} does not reference D{r}*E{r}: {f_val}")

        # G should reference B{r}
        g_val = get_formula(ws, f"G{r}")
        if f"B{r}" not in g_val:
            row_errors.append(f"G{r} does not reference B{r}: {g_val}")

        # H should reference C{r}
        h_val = get_formula(ws, f"H{r}")
        if f"C{r}" not in h_val:
            row_errors.append(f"H{r} does not reference C{r}: {h_val}")

        # I should reference F{r}
        i_val = get_formula(ws, f"I{r}")
        if f"F{r}" not in i_val:
            row_errors.append(f"I{r} does not reference F{r}: {i_val}")

        # J should reference E{r}
        j_val = get_formula(ws, f"J{r}")
        if f"E{r}" not in j_val:
            row_errors.append(f"J{r} does not reference E{r}: {j_val}")

        # K should reference G{r}, H{r}, I{r}, J{r}
        k_val = get_formula(ws, f"K{r}")
        for col in ["G", "H", "I", "J"]:
            if f"{col}{r}" not in k_val:
                row_errors.append(f"K{r} does not reference {col}{r}: {k_val}")

        if row_errors:
            all_rows_ok = False
            for e in row_errors:
                log_error(e)
        else:
            log_pass(f"Row {r}: all formulas reference correct row")

    if all_rows_ok:
        log_pass("Row-reference integrity: ALL sampled rows correct")


# ---------------------------------------------------------------------------
# Price History Sheet Audit
# ---------------------------------------------------------------------------
def audit_price_history(wb):
    print("\n=== Auditing Price History Sheet ===")
    ws = wb["Price History"]

    # Check block structure: 100 blocks at 320-row intervals starting at row 3
    blocks_to_check = [0, 1, 49, 99]

    for i in blocks_to_check:
        start_row = 3 + i * 320
        ticker_input_row = 3 + i

        print(f"\n  Checking block {i} (start_row={start_row}, ticker_input_row={ticker_input_row})")

        # A cell should reference ticker input column I
        a_val = get_formula(ws, f"A{start_row}")
        expected_a = f"=I{ticker_input_row}"
        if a_val == expected_a:
            log_pass(f"Block {i}: A{start_row} = {expected_a}")
        else:
            log_error(f"Block {i}: A{start_row} expected {expected_a}, got: {a_val}")

        # B cell: BDH formula checks
        b_val = get_formula(ws, f"B{start_row}")

        # Check fields
        if "PX_OPEN,PX_HIGH,PX_LOW,PX_LAST,VOLUME" in b_val:
            log_pass(f"Block {i}: B{start_row} has correct OHLCV fields")
        else:
            log_error(f"Block {i}: B{start_row} missing OHLCV fields: {b_val}")

        # Check start date
        if "EDATE(TODAY(),-15)" in b_val:
            log_pass(f"Block {i}: B{start_row} has dynamic 15-month lookback")
        else:
            log_error(f"Block {i}: B{start_row} missing EDATE(TODAY(),-15): {b_val}")

        # Check corporate action adjustments
        if 'CshAdjNormal","Y"' in b_val:
            log_pass(f"Block {i}: B{start_row} has CshAdjNormal=Y")
        else:
            log_error(f"Block {i}: B{start_row} missing CshAdjNormal: {b_val}")

        if 'CshAdjAbnormal","Y"' in b_val:
            log_pass(f"Block {i}: B{start_row} has CshAdjAbnormal=Y")
        else:
            log_error(f"Block {i}: B{start_row} missing CshAdjAbnormal: {b_val}")

        # Check that BDH references A{start_row}
        if f"A{start_row}" in b_val:
            log_pass(f"Block {i}: B{start_row} references A{start_row}")
        else:
            log_error(f"Block {i}: B{start_row} does not reference A{start_row}: {b_val}")

    # Verify blocks exist for all 100 tickers
    print("\n  Verifying all 100 blocks exist...")
    missing_blocks = []
    for i in range(100):
        start_row = 3 + i * 320
        a_val = get_formula(ws, f"A{start_row}")
        if not a_val or not a_val.startswith("="):
            missing_blocks.append(i)

    if missing_blocks:
        log_error(f"Missing blocks: {missing_blocks}")
    else:
        log_pass("All 100 Price History blocks exist")


# ---------------------------------------------------------------------------
# Fundamentals Sheet Audit
# ---------------------------------------------------------------------------
def audit_fundamentals(wb):
    print("\n=== Auditing Fundamentals Sheet ===")
    ws = wb["Fundamentals"]

    # Check blocks 0, 1, 49, 99
    blocks_to_check = [0, 1, 49, 99]

    for i in blocks_to_check:
        sr = 2 + i * 10  # start row
        ticker_input_row = 2 + i

        print(f"\n  Checking block {i} (start_row={sr}, ticker_input_row={ticker_input_row})")

        # A cell references ticker input column M
        a_val = get_formula(ws, f"A{sr}")
        expected_a = f"=M{ticker_input_row}"
        if a_val == expected_a:
            log_pass(f"Block {i}: A{sr} = {expected_a}")
        else:
            log_error(f"Block {i}: A{sr} expected {expected_a}, got: {a_val}")

        # B cell: EPS BDH
        b_val = get_formula(ws, f"B{sr}")
        if "IS_DILUTED_EPS" in b_val:
            log_pass(f"Block {i}: B{sr} has IS_DILUTED_EPS")
        else:
            log_error(f"Block {i}: B{sr} missing IS_DILUTED_EPS: {b_val}")

        if "EDATE(TODAY(),-24)" in b_val:
            log_pass(f"Block {i}: B{sr} has 24-month lookback")
        else:
            log_error(f"Block {i}: B{sr} missing 24-month lookback: {b_val}")

        if f"A{sr}" in b_val:
            log_pass(f"Block {i}: B{sr} references A{sr}")
        else:
            log_error(f"Block {i}: B{sr} does not reference A{sr}: {b_val}")

        # J cell: Revenue BDH
        j_val = get_formula(ws, f"J{sr}")
        if "SALES_REV_TURN" in j_val:
            log_pass(f"Block {i}: J{sr} has SALES_REV_TURN")
        else:
            log_error(f"Block {i}: J{sr} missing SALES_REV_TURN: {j_val}")

        if "EDATE(TODAY(),-24)" in j_val:
            log_pass(f"Block {i}: J{sr} has 24-month lookback")
        else:
            log_error(f"Block {i}: J{sr} missing 24-month lookback: {j_val}")

        # L cell: EPS YoY growth formula
        l_val = get_formula(ws, f"L{sr}")
        eps_newest = f"C{sr+7}"
        eps_old = f"C{sr+3}"

        if eps_newest in l_val and eps_old in l_val:
            log_pass(f"Block {i}: L{sr} references {eps_newest} (newest) and {eps_old} (4Q ago)")
        else:
            log_error(f"Block {i}: L{sr} missing correct cell refs ({eps_newest}/{eps_old}): {l_val}")

        # Check for zero/negative denominator guards
        if "IF(" in l_val and ("=0" in l_val or "<=0" in l_val or "<0" in l_val):
            log_pass(f"Block {i}: L{sr} has IF guard for denominator")
        else:
            log_error(f"Block {i}: L{sr} missing IF guard for zero/negative denominator: {l_val}")

        # Check for 999 turnaround value
        if "999" in l_val:
            log_pass(f"Block {i}: L{sr} uses 999 for turnaround")
        else:
            log_error(f"Block {i}: L{sr} missing 999 turnaround value: {l_val}")

        # N cell: Revenue YoY growth formula
        n_val = get_formula(ws, f"N{sr}")
        rev_newest = f"K{sr+7}"
        rev_old = f"K{sr+3}"

        if rev_newest in n_val and rev_old in n_val:
            log_pass(f"Block {i}: N{sr} references {rev_newest} (newest) and {rev_old} (4Q ago)")
        else:
            log_error(f"Block {i}: N{sr} missing correct cell refs ({rev_newest}/{rev_old}): {n_val}")

        # Check for denominator guard in N
        if "IF(" in n_val and ("<=0" in n_val or "=0" in n_val):
            log_pass(f"Block {i}: N{sr} has IF guard for denominator")
        else:
            log_error(f"Block {i}: N{sr} missing IF guard for zero/negative denominator: {n_val}")

        if "999" in n_val:
            log_pass(f"Block {i}: N{sr} uses 999 for turnaround")
        else:
            log_error(f"Block {i}: N{sr} missing 999 turnaround value: {n_val}")

    # Verify all 100 blocks exist
    print("\n  Verifying all 100 Fundamentals blocks exist...")
    missing_blocks = []
    for i in range(100):
        sr = 2 + i * 10
        a_val = get_formula(ws, f"A{sr}")
        if not a_val or not a_val.startswith("="):
            missing_blocks.append(i)

    if missing_blocks:
        log_error(f"Missing Fundamentals blocks: {missing_blocks}")
    else:
        log_pass("All 100 Fundamentals blocks exist")


# ---------------------------------------------------------------------------
# Factor Scores Sheet Audit
# ---------------------------------------------------------------------------
def audit_factor_scores(wb):
    print("\n=== Auditing Factor Scores Sheet ===")
    ws = wb["Factor Scores"]

    # Check row 2 formulas
    a2 = get_formula(ws, "A2")
    if a2 == "=Universe!A2":
        log_pass("A2: References Universe!A2 correctly")
    else:
        log_error(f"A2: Expected =Universe!A2, got: {a2}")

    c2 = get_formula(ws, "C2")
    if "PERCENTRANK.INC" in c2:
        log_pass("C2: Uses PERCENTRANK.INC")
    else:
        log_error(f"C2: Expected PERCENTRANK.INC, got: {c2}")

    e2 = get_formula(ws, "E2")
    if "MIN(D2,100)" in e2:
        log_pass("E2: Uses MIN(D2,100)")
    else:
        log_error(f"E2: Expected MIN(D2,100), got: {e2}")

    g2 = get_formula(ws, "G2")
    if "MIN(F2,100)" in g2:
        log_pass("G2: Uses MIN(F2,100)")
    else:
        log_error(f"G2: Expected MIN(F2,100), got: {g2}")

    h2 = get_formula(ws, "H2")
    if "BDP" in h2 and "HIGH_52WEEK" in h2:
        log_pass("H2: Uses BDP with HIGH_52WEEK")
    else:
        log_error(f"H2: Expected BDP with HIGH_52WEEK, got: {h2}")

    # Quality gate columns
    i2 = get_formula(ws, "I2")
    if "D2>=5" in i2:
        log_pass("I2: Quality gate D2>=5 correct")
    else:
        log_error(f"I2: Expected D2>=5, got: {i2}")

    j2 = get_formula(ws, "J2")
    if "F2>=5" in j2:
        log_pass("J2: Quality gate F2>=5 correct")
    else:
        log_error(f"J2: Expected F2>=5, got: {j2}")

    k2 = get_formula(ws, "K2")
    if "H2>=75" in k2:
        log_pass("K2: Quality gate H2>=75 correct")
    else:
        log_error(f"K2: Expected H2>=75, got: {k2}")

    l2 = get_formula(ws, "L2")
    if "AND(I2,J2,K2)" in l2 or "AND(I2, J2, K2)" in l2:
        log_pass("L2: AND of quality gates correct")
    else:
        log_error(f"L2: Expected AND(I2,J2,K2), got: {l2}")

    # Check a few random rows for correct row references
    random.seed(99)
    sample_rows = random.sample(range(3, 2502), 5)
    print(f"\n  Checking Factor Scores row-reference integrity: {sorted(sample_rows)}")

    for r in sorted(sample_rows):
        row_errors = []

        a_val = get_formula(ws, f"A{r}")
        if f"Universe!A{r}" not in a_val:
            row_errors.append(f"A{r} does not reference Universe!A{r}: {a_val}")

        c_val = get_formula(ws, f"C{r}")
        if f"B{r}" not in c_val:
            row_errors.append(f"C{r} does not reference B{r}: {c_val}")

        e_val = get_formula(ws, f"E{r}")
        if f"D{r}" not in e_val:
            row_errors.append(f"E{r} does not reference D{r}: {e_val}")

        g_val = get_formula(ws, f"G{r}")
        if f"F{r}" not in g_val:
            row_errors.append(f"G{r} does not reference F{r}: {g_val}")

        h_val = get_formula(ws, f"H{r}")
        if f"A{r}" not in h_val:
            row_errors.append(f"H{r} does not reference A{r}: {h_val}")
        if f"Universe!E{r}" not in h_val:
            row_errors.append(f"H{r} does not reference Universe!E{r}: {h_val}")

        i_val = get_formula(ws, f"I{r}")
        if f"D{r}" not in i_val:
            row_errors.append(f"I{r} does not reference D{r}: {i_val}")

        j_val = get_formula(ws, f"J{r}")
        if f"F{r}" not in j_val:
            row_errors.append(f"J{r} does not reference F{r}: {j_val}")

        k_val = get_formula(ws, f"K{r}")
        if f"H{r}" not in k_val:
            row_errors.append(f"K{r} does not reference H{r}: {k_val}")

        l_val = get_formula(ws, f"L{r}")
        for col in ["I", "J", "K"]:
            if f"{col}{r}" not in l_val:
                row_errors.append(f"L{r} does not reference {col}{r}: {l_val}")

        if row_errors:
            for e in row_errors:
                log_error(e)
        else:
            log_pass(f"Row {r}: all Factor Scores formulas reference correct row")


# ---------------------------------------------------------------------------
# Portfolio Sheet Audit
# ---------------------------------------------------------------------------
def audit_portfolio(wb):
    print("\n=== Auditing Portfolio Sheet ===")
    ws = wb["Portfolio"]

    # A2
    a2 = get_formula(ws, "A2")
    if "'Factor Scores'!A2" in a2:
        log_pass("A2: References Factor Scores!A2")
    else:
        log_error(f"A2: Expected 'Factor Scores'!A2 reference, got: {a2}")

    # B2: composite formula with weights
    b2 = get_formula(ws, "B2")
    if "0.40" in b2 or "0.4" in b2:
        log_pass("B2: Has 0.40 weight")
    else:
        log_error(f"B2: Missing 0.40 weight: {b2}")

    # Check correct column references in B2
    if "'Factor Scores'!C2" in b2:
        log_pass("B2: References Factor Scores C (RS Percentile)")
    else:
        log_error(f"B2: Missing Factor Scores C reference: {b2}")

    if "'Factor Scores'!E2" in b2:
        log_pass("B2: References Factor Scores E (EPS capped)")
    else:
        log_error(f"B2: Missing Factor Scores E reference: {b2}")

    if "'Factor Scores'!G2" in b2:
        log_pass("B2: References Factor Scores G (Rev capped)")
    else:
        log_error(f"B2: Missing Factor Scores G reference: {b2}")

    if "'Factor Scores'!H2" in b2:
        log_pass("B2: References Factor Scores H (52Wk High %)")
    else:
        log_error(f"B2: Missing Factor Scores H reference: {b2}")

    # Check weight distribution: 0.40 + 0.20 + 0.20 + 0.20 = 1.00
    weight_count = b2.count("0.20")
    if weight_count == 3:
        log_pass("B2: Has three 0.20 weights")
    else:
        log_error(f"B2: Expected three 0.20 weights, found {weight_count}: {b2}")

    # C2: RANK formula
    c2 = get_formula(ws, "C2")
    if "RANK" in c2 and "B2" in c2:
        log_pass("C2: RANK formula correct")
    else:
        log_error(f"C2: Expected RANK with B2, got: {c2}")

    # D2: Top 25 check
    d2 = get_formula(ws, "D2")
    if "C2<=25" in d2:
        log_pass("D2: C2<=25 check correct")
    else:
        log_error(f"D2: Expected C2<=25, got: {d2}")

    # F2: Signal logic
    f2 = get_formula(ws, "F2")
    if "BUY" in f2 and "SELL" in f2 and "HOLD" in f2:
        log_pass("F2: Signal logic has BUY/SELL/HOLD")
    else:
        log_error(f"F2: Missing BUY/SELL/HOLD logic: {f2}")

    if "D2" in f2 and "E2" in f2:
        log_pass("F2: References D2 (Top25) and E2 (Prior Month)")
    else:
        log_error(f"F2: Missing D2/E2 references: {f2}")

    # H2: BDP with GICS_SECTOR_NAME
    h2 = get_formula(ws, "H2")
    if "BDP" in h2 and "GICS_SECTOR_NAME" in h2:
        log_pass("H2: BDP with GICS_SECTOR_NAME correct")
    else:
        log_error(f"H2: Expected BDP with GICS_SECTOR_NAME, got: {h2}")

    # Check a few random rows for correct references
    random.seed(77)
    sample_rows = random.sample(range(3, 2502), 5)
    print(f"\n  Checking Portfolio row-reference integrity: {sorted(sample_rows)}")

    for r in sorted(sample_rows):
        row_errors = []

        a_val = get_formula(ws, f"A{r}")
        if f"'Factor Scores'!A{r}" not in a_val:
            row_errors.append(f"A{r} does not reference 'Factor Scores'!A{r}: {a_val}")

        b_val = get_formula(ws, f"B{r}")
        for col in ["C", "E", "G", "H"]:
            if f"'Factor Scores'!{col}{r}" not in b_val:
                row_errors.append(f"B{r} does not reference 'Factor Scores'!{col}{r}: {b_val}")

        c_val = get_formula(ws, f"C{r}")
        if f"B{r}" not in c_val:
            row_errors.append(f"C{r} does not reference B{r}: {c_val}")

        d_val = get_formula(ws, f"D{r}")
        if f"C{r}" not in d_val:
            row_errors.append(f"D{r} does not reference C{r}: {d_val}")

        f_val = get_formula(ws, f"F{r}")
        if f"D{r}" not in f_val or f"E{r}" not in f_val:
            row_errors.append(f"F{r} does not reference D{r}/E{r}: {f_val}")

        h_val = get_formula(ws, f"H{r}")
        if f"A{r}" not in h_val:
            row_errors.append(f"H{r} does not reference A{r}: {h_val}")

        if row_errors:
            for e in row_errors:
                log_error(e)
        else:
            log_pass(f"Row {r}: all Portfolio formulas reference correct row")


# ---------------------------------------------------------------------------
# Deep formula correctness checks
# ---------------------------------------------------------------------------
def audit_deep_checks(wb):
    """Additional cross-cutting checks."""
    print("\n=== Deep Cross-Sheet Checks ===")

    # Check that EPS YoY in Fundamentals handles the edge case where
    # denominator is negative properly (should assign 999 for neg->pos turnaround)
    ws = wb["Fundamentals"]
    l2 = get_formula(ws, "L2")

    # The EPS formula should handle: denominator=0 AND denominator<0
    if "=0" in l2 and "<0" in l2:
        log_pass("L2 EPS: Handles both zero AND negative denominators")
    elif "<=0" in l2:
        log_pass("L2 EPS: Handles <=0 denominator (combined guard)")
    elif "=0" in l2 and "<0" not in l2:
        log_error("L2 EPS: Handles zero but NOT negative denominators separately")
    else:
        log_error(f"L2 EPS: Denominator guard unclear: {l2}")

    # Check that Revenue YoY handles negative denominators
    n2 = get_formula(ws, "N2")
    if "<=0" in n2:
        log_pass("N2 Rev: Handles <=0 denominator (combined guard)")
    elif "=0" in n2 and "<0" in n2:
        log_pass("N2 Rev: Handles both zero AND negative denominators")
    elif "=0" in n2:
        log_warn("N2 Rev: Only handles zero denominator, not negative")
    else:
        log_error(f"N2 Rev: Denominator guard unclear: {n2}")

    # Verify EPS uses ABS() in denominator to handle negative base EPS correctly
    if "ABS(" in l2:
        log_pass("L2 EPS: Uses ABS() in denominator for correct growth calc from negative base")
    else:
        log_warn("L2 EPS: Does not use ABS() in denominator — growth % from negative base will be inverted")

    # Check that Revenue formula also uses proper division
    if "ABS(" in n2 or "<=0" in n2:
        # If it catches <=0, it routes around the division entirely
        log_pass("N2 Rev: Division protected (either ABS or guard)")
    else:
        log_warn(f"N2 Rev: May not handle negative denominator in division: {n2}")

    # Verify Factor Scores H column formula structure
    ws_fs = wb["Factor Scores"]
    h2 = get_formula(ws_fs, "H2")
    if "Universe!E2" in h2 and "BDP" in h2 and "HIGH_52WEEK" in h2:
        log_pass("Factor Scores H2: (PX_LAST / 52WkHigh) * 100 structure correct")
    else:
        log_error(f"Factor Scores H2: Structure issue: {h2}")

    # Verify Portfolio composite weights sum to 1.0
    ws_p = wb["Portfolio"]
    b2 = get_formula(ws_p, "B2")
    # Extract weights
    import re
    weights = re.findall(r'(\d+\.\d+)\*', b2)
    if weights:
        total = sum(float(w) for w in weights)
        if abs(total - 1.0) < 0.001:
            log_pass(f"Portfolio B2: Weights sum to {total:.2f} (correct)")
        else:
            log_error(f"Portfolio B2: Weights sum to {total:.2f} (should be 1.00)")
    else:
        log_error(f"Portfolio B2: Could not extract weights from: {b2}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"Loading template: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)

    print(f"Sheets found: {wb.sheetnames}")
    expected_sheets = ["README", "Universe", "Price History", "Fundamentals", "Factor Scores", "Portfolio"]
    for name in expected_sheets:
        if name in wb.sheetnames:
            log_pass(f"Sheet '{name}' exists")
        else:
            log_error(f"Sheet '{name}' MISSING")

    audit_universe(wb)
    audit_price_history(wb)
    audit_fundamentals(wb)
    audit_factor_scores(wb)
    audit_portfolio(wb)
    audit_deep_checks(wb)

    # Summary
    print("\n" + "=" * 60)
    print("AUDIT SUMMARY")
    print("=" * 60)
    print(f"  PASS:    {len(passes)}")
    print(f"  WARN:    {len(warnings)}")
    print(f"  ERROR:   {len(errors)}")

    if warnings:
        print("\nWARNINGS:")
        for w in warnings:
            print(f"  - {w}")

    if errors:
        print("\nERRORS:")
        for e in errors:
            print(f"  - {e}")
        print("\nAUDIT RESULT: FAIL")
        return 1
    else:
        print("\nAUDIT RESULT: PASS")
        return 0


if __name__ == "__main__":
    sys.exit(main())

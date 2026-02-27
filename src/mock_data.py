"""Generate realistic mock Bloomberg Excel workbooks for testing.

Creates a 5-sheet workbook matching the Bloomberg template structure with
simulated small-cap universe data, price histories (GBM), and quarterly
fundamentals (EPS + revenue).

Usage:
    from src.mock_data import generate_mock_data
    path = generate_mock_data("tests/fixtures/mock_bloomberg.xlsx")
"""

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_EXCHANGES = ["NAS", "NYS", "ASE"]
_EXCHANGE_WEIGHTS = [0.50, 0.35, 0.15]  # distribution across exchanges

_PRICE_DAYS = 315  # ~15 months of trading days
_PRICE_BLOCK_SIZE = 320
_FUND_QUARTERS = 8
_FUND_BLOCK_SIZE = 10


# ---------------------------------------------------------------------------
# Helper: geometric Brownian motion price paths
# ---------------------------------------------------------------------------


def _gbm_prices(
    rng: np.random.Generator,
    start_price: float,
    n_days: int,
    mu: float,
    sigma: float,
) -> np.ndarray:
    """Simulate daily close prices via geometric Brownian motion.

    Args:
        rng: numpy random generator
        start_price: initial price
        n_days: number of trading days
        mu: daily drift (e.g. 0.0005 = 0.05%/day)
        sigma: daily volatility (e.g. 0.02 = 2%/day)

    Returns:
        Array of shape (n_days,) with simulated close prices.
    """
    dt = 1.0
    log_returns = (mu - 0.5 * sigma**2) * dt + sigma * np.sqrt(dt) * rng.standard_normal(n_days)
    prices = start_price * np.exp(np.cumsum(log_returns))
    return prices


def _generate_ohlcv(
    rng: np.random.Generator,
    closes: np.ndarray,
    avg_volume: float,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Generate OHLV data from a close-price series.

    Returns (open, high, low, close, volume) arrays.
    """
    n = len(closes)

    # Open: previous close + small gap
    opens = np.empty(n)
    opens[0] = closes[0] * (1 + rng.normal(0, 0.003))
    opens[1:] = closes[:-1] * (1 + rng.normal(0, 0.003, n - 1))

    # High / Low: extend beyond max/min of open/close
    day_max = np.maximum(opens, closes)
    day_min = np.minimum(opens, closes)
    highs = day_max * (1 + np.abs(rng.normal(0, 0.005, n)))
    lows = day_min * (1 - np.abs(rng.normal(0, 0.005, n)))

    # Ensure lows > 0
    lows = np.maximum(lows, 0.01)

    # Volume: log-normal around avg_volume
    volumes = rng.lognormal(
        mean=np.log(avg_volume) - 0.5,  # median near avg_volume
        sigma=0.5,
        size=n,
    ).astype(int)

    return opens, highs, lows, closes, volumes


# ---------------------------------------------------------------------------
# Sheet generators
# ---------------------------------------------------------------------------


def _generate_universe(
    rng: np.random.Generator,
    tickers: list[str],
    ticker_profiles: list[dict],
) -> pd.DataFrame:
    """Build the Universe sheet DataFrame (columns A-K, we only care about A-F)."""
    rows = []
    for i, (tk, prof) in enumerate(zip(tickers, ticker_profiles)):
        rows.append(
            {
                "Ticker": tk,
                "Market Cap ($M)": prof["market_cap"],
                "Exchange": prof["exchange"],
                "Avg Volume (20d)": prof["avg_volume"],
                "Price": prof["price"],
                "Dollar Volume": prof["avg_volume"] * prof["price"],
            }
        )
    return pd.DataFrame(rows)


def _write_price_history_sheet(
    ws,
    tickers: list[str],
    ticker_profiles: list[dict],
    rng: np.random.Generator,
):
    """Write the Price History sheet in block layout (320 rows per ticker)."""
    # Generate a date index: 315 trading days ending ~today
    end_date = pd.Timestamp("2026-02-20")
    dates = pd.bdate_range(end=end_date, periods=_PRICE_DAYS)

    current_row = 1  # openpyxl is 1-indexed

    for tk, prof in zip(tickers, ticker_profiles):
        block_start = current_row

        # Row 1 of block: ticker in column A
        ws.cell(row=current_row, column=1, value=tk)
        current_row += 1

        # Generate price path
        closes = _gbm_prices(
            rng,
            start_price=prof["start_price"],
            n_days=_PRICE_DAYS,
            mu=prof["drift"],
            sigma=prof["vol"],
        )

        opens, highs, lows, closes, volumes = _generate_ohlcv(rng, closes, prof["avg_volume"])

        # Update profile with ending price (used for universe sheet "Price" column)
        prof["price"] = round(float(closes[-1]), 2)

        # Introduce a few #N/A values for realism (~0.3% of cells)
        na_mask = rng.random(_PRICE_DAYS) < 0.003

        for j in range(_PRICE_DAYS):
            dt = dates[j]
            if na_mask[j]:
                # Write Bloomberg-style #N/A for this row
                ws.cell(row=current_row, column=2, value=dt.strftime("%Y-%m-%d"))
                ws.cell(row=current_row, column=3, value="#N/A")
                ws.cell(row=current_row, column=4, value="#N/A")
                ws.cell(row=current_row, column=5, value="#N/A")
                ws.cell(row=current_row, column=6, value="#N/A")
                ws.cell(row=current_row, column=7, value="#N/A")
            else:
                ws.cell(row=current_row, column=2, value=dt.strftime("%Y-%m-%d"))
                ws.cell(row=current_row, column=3, value=round(float(opens[j]), 4))
                ws.cell(row=current_row, column=4, value=round(float(highs[j]), 4))
                ws.cell(row=current_row, column=5, value=round(float(lows[j]), 4))
                ws.cell(row=current_row, column=6, value=round(float(closes[j]), 4))
                ws.cell(row=current_row, column=7, value=int(volumes[j]))
            current_row += 1

        # Pad remaining rows in the 320-row block with blanks
        rows_written = current_row - block_start
        padding = _PRICE_BLOCK_SIZE - rows_written
        current_row += padding  # skip blank rows


def _write_fundamentals_sheet(
    ws,
    tickers: list[str],
    ticker_profiles: list[dict],
    rng: np.random.Generator,
):
    """Write the Fundamentals sheet in block layout (10 rows per ticker)."""
    # 8 quarter-end dates going back 2 years
    quarter_dates = pd.date_range(end="2025-12-31", periods=_FUND_QUARTERS, freq="QE")

    current_row = 1

    for tk, prof in zip(tickers, ticker_profiles):
        # Row 1 of block: ticker in column A
        ws.cell(row=current_row, column=1, value=tk)
        current_row += 1

        # Generate EPS trajectory
        eps_base = prof["eps_base"]
        eps_growth = prof["eps_growth"]  # quarterly YoY growth rate
        eps_values = []
        for q in range(_FUND_QUARTERS):
            # Growth compounds: quarter q has q quarters of growth applied
            if prof["turnaround"] and q < 4:
                # First 4 quarters negative, then turn positive
                val = eps_base * (1 + eps_growth * q) - abs(eps_base) * 1.5
            else:
                val = eps_base * (1 + eps_growth * q)
            val += rng.normal(0, abs(eps_base) * 0.05)  # small noise
            eps_values.append(round(val, 2))

        # Generate revenue trajectory
        rev_base = prof["rev_base"]  # quarterly revenue in $M
        rev_growth = prof["rev_growth"]
        rev_values = []
        for q in range(_FUND_QUARTERS):
            val = rev_base * (1 + rev_growth * q)
            val += rng.normal(0, rev_base * 0.03)
            rev_values.append(round(max(val, 0), 1))

        # Decide if some quarters should be missing
        n_quarters_to_write = _FUND_QUARTERS
        if prof.get("missing_quarters", False):
            n_quarters_to_write = rng.integers(4, 7)  # write 4-6 quarters instead of 8

        for q in range(n_quarters_to_write):
            dt_str = quarter_dates[q].strftime("%Y-%m-%d")

            # Handle pre-revenue tickers
            if prof.get("pre_revenue", False) and q < 5:
                rev_val = "#N/A"
            else:
                rev_val = rev_values[q]

            # EPS: columns B-C
            ws.cell(row=current_row, column=2, value=dt_str)
            ws.cell(row=current_row, column=3, value=eps_values[q])

            # Revenue: columns E-F (skip column D)
            ws.cell(row=current_row, column=5, value=dt_str)
            ws.cell(row=current_row, column=6, value=rev_val)

            current_row += 1

        # Pad to 10-row block
        rows_written = 1 + n_quarters_to_write  # 1 for ticker row
        padding = _FUND_BLOCK_SIZE - rows_written
        current_row += padding


# ---------------------------------------------------------------------------
# Profile generation: assign characteristics to each ticker
# ---------------------------------------------------------------------------


def _generate_profiles(
    rng: np.random.Generator,
    n_tickers: int,
) -> list[dict]:
    """Create ticker profiles that produce a realistic distribution:
    - ~70 pass universe filters (market cap, price, volume, exchange)
    - ~25-35 pass quality filters (EPS growth, revenue growth)
    - Clear top-25 candidates with strong RS + fundamentals
    - Edge cases: turnarounds, pre-revenue, penny stocks, OTC
    """
    profiles = []

    for i in range(n_tickers):
        prof: dict = {}

        # --- Archetype assignment ---
        # ~30% strong bullish (clear winners)
        # ~25% moderate / mixed
        # ~25% weak / declining
        # ~10% edge cases (turnaround, pre-revenue, penny, OTC)
        # ~10% filter-fail (too small, OTC, penny)

        r = i / n_tickers  # deterministic bucketing for reproducibility

        if r < 0.30:
            # --- STRONG: high growth, bullish price action ---
            prof["archetype"] = "strong"
            prof["market_cap"] = rng.uniform(200, 5000)
            prof["exchange"] = rng.choice(["NAS", "NYS"], p=[0.6, 0.4])
            prof["avg_volume"] = rng.uniform(500_000, 4_000_000)
            prof["start_price"] = rng.uniform(15, 80)
            prof["drift"] = rng.uniform(0.0008, 0.002)  # bullish
            prof["vol"] = rng.uniform(0.015, 0.03)
            prof["eps_base"] = rng.uniform(0.30, 2.50)
            prof["eps_growth"] = rng.uniform(0.06, 0.15)  # strong YoY
            prof["rev_base"] = rng.uniform(30, 400)
            prof["rev_growth"] = rng.uniform(0.06, 0.12)
            prof["turnaround"] = False

        elif r < 0.55:
            # --- MODERATE: decent but not standout ---
            prof["archetype"] = "moderate"
            prof["market_cap"] = rng.uniform(100, 3000)
            prof["exchange"] = rng.choice(_EXCHANGES, p=_EXCHANGE_WEIGHTS)
            prof["avg_volume"] = rng.uniform(300_000, 2_000_000)
            prof["start_price"] = rng.uniform(8, 60)
            prof["drift"] = rng.uniform(-0.0002, 0.0005)
            prof["vol"] = rng.uniform(0.018, 0.035)
            prof["eps_base"] = rng.uniform(0.10, 1.00)
            prof["eps_growth"] = rng.uniform(0.01, 0.06)
            prof["rev_base"] = rng.uniform(20, 250)
            prof["rev_growth"] = rng.uniform(0.02, 0.06)
            prof["turnaround"] = False

        elif r < 0.80:
            # --- WEAK: declining or flat ---
            prof["archetype"] = "weak"
            prof["market_cap"] = rng.uniform(80, 2000)
            prof["exchange"] = rng.choice(_EXCHANGES, p=_EXCHANGE_WEIGHTS)
            prof["avg_volume"] = rng.uniform(200_000, 1_500_000)
            prof["start_price"] = rng.uniform(5, 40)
            prof["drift"] = rng.uniform(-0.001, -0.0002)
            prof["vol"] = rng.uniform(0.02, 0.04)
            prof["eps_base"] = rng.uniform(-0.20, 0.50)
            prof["eps_growth"] = rng.uniform(-0.05, 0.02)
            prof["rev_base"] = rng.uniform(15, 200)
            prof["rev_growth"] = rng.uniform(-0.02, 0.03)
            prof["turnaround"] = False

        elif r < 0.90:
            # --- EDGE CASES: turnarounds, pre-revenue ---
            edge_type = rng.choice(["turnaround", "pre_revenue", "missing_data"])
            prof["archetype"] = edge_type

            prof["market_cap"] = rng.uniform(100, 1500)
            prof["exchange"] = rng.choice(["NAS", "NYS"], p=[0.7, 0.3])
            prof["avg_volume"] = rng.uniform(400_000, 2_000_000)
            prof["start_price"] = rng.uniform(5, 30)
            prof["drift"] = rng.uniform(0.0003, 0.001)
            prof["vol"] = rng.uniform(0.025, 0.045)

            if edge_type == "turnaround":
                prof["eps_base"] = rng.uniform(0.20, 0.80)
                prof["eps_growth"] = rng.uniform(0.08, 0.15)
                prof["turnaround"] = True  # neg->pos EPS pattern
                prof["rev_base"] = rng.uniform(20, 150)
                prof["rev_growth"] = rng.uniform(0.04, 0.10)
            elif edge_type == "pre_revenue":
                prof["eps_base"] = rng.uniform(-0.50, -0.10)
                prof["eps_growth"] = rng.uniform(0.02, 0.05)
                prof["turnaround"] = False
                prof["pre_revenue"] = True
                prof["rev_base"] = rng.uniform(5, 20)
                prof["rev_growth"] = rng.uniform(0.10, 0.25)
            else:  # missing_data
                prof["eps_base"] = rng.uniform(0.10, 0.60)
                prof["eps_growth"] = rng.uniform(0.03, 0.08)
                prof["turnaround"] = False
                prof["missing_quarters"] = True
                prof["rev_base"] = rng.uniform(15, 100)
                prof["rev_growth"] = rng.uniform(0.03, 0.07)

        else:
            # --- FILTER FAILURES: penny stocks, OTC, tiny market cap ---
            fail_type = rng.choice(["penny", "otc", "tiny_cap", "low_volume"])
            prof["archetype"] = f"fail_{fail_type}"

            if fail_type == "penny":
                prof["market_cap"] = rng.uniform(30, 200)
                prof["exchange"] = rng.choice(["NAS", "ASE"])
                prof["avg_volume"] = rng.uniform(100_000, 800_000)
                prof["start_price"] = rng.uniform(0.50, 1.80)  # below $2
            elif fail_type == "otc":
                prof["market_cap"] = rng.uniform(50, 500)
                prof["exchange"] = "OTC"
                prof["avg_volume"] = rng.uniform(100_000, 500_000)
                prof["start_price"] = rng.uniform(3, 20)
            elif fail_type == "tiny_cap":
                prof["market_cap"] = rng.uniform(10, 45)  # below $50M
                prof["exchange"] = rng.choice(["NAS", "NYS"])
                prof["avg_volume"] = rng.uniform(50_000, 300_000)
                prof["start_price"] = rng.uniform(3, 15)
            else:  # low_volume
                prof["market_cap"] = rng.uniform(100, 800)
                prof["exchange"] = rng.choice(["NAS", "ASE"])
                prof["avg_volume"] = rng.uniform(10_000, 80_000)  # very thin
                prof["start_price"] = rng.uniform(5, 30)

            prof["drift"] = rng.uniform(-0.001, 0.0005)
            prof["vol"] = rng.uniform(0.03, 0.06)
            prof["eps_base"] = rng.uniform(-0.30, 0.30)
            prof["eps_growth"] = rng.uniform(-0.03, 0.03)
            prof["rev_base"] = rng.uniform(5, 50)
            prof["rev_growth"] = rng.uniform(-0.02, 0.04)
            prof["turnaround"] = False

        # Default flags
        prof.setdefault("pre_revenue", False)
        prof.setdefault("missing_quarters", False)
        prof.setdefault("turnaround", False)

        # Price will be updated after GBM simulation (ending price)
        prof["price"] = round(prof["start_price"], 2)

        profiles.append(prof)

    return profiles


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def generate_mock_data(
    output_path: str | Path,
    n_tickers: int = 100,
    seed: int = 42,
) -> Path:
    """Generate a realistic mock Bloomberg Excel workbook.

    Creates a 5-sheet workbook matching the Bloomberg template structure:

    Sheet 1 'Universe': fake tickers with realistic small-cap data
    Sheet 2 'Price History': 315 days of simulated OHLCV per ticker (GBM)
    Sheet 3 'Fundamentals': 8 quarters of EPS and revenue per ticker
    Sheets 4-5: Empty placeholders (engine computes these)

    Args:
        output_path: Where to write the .xlsx file
        n_tickers: Number of tickers to generate (default 100)
        seed: Random seed for reproducibility (default 42)

    Returns:
        Path to the created file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rng = np.random.default_rng(seed)
    logger.info("Generating mock data: %d tickers, seed=%d", n_tickers, seed)

    # Ticker symbols
    tickers = [f"MOCK{i+1:03d}" for i in range(n_tickers)]

    # Generate profiles
    profiles = _generate_profiles(rng, n_tickers)

    # --- Build workbook ---
    wb = Workbook()

    # Sheet 1: Universe (created as default active sheet)
    ws_universe = wb.active
    ws_universe.title = "Universe"

    # Write header
    headers = ["Ticker", "Market Cap ($M)", "Exchange", "Avg Volume (20d)", "Price", "Dollar Volume"]
    for col_idx, h in enumerate(headers, start=1):
        ws_universe.cell(row=1, column=col_idx, value=h)

    # We need to write price history first so that profiles get their ending prices
    # Then we can write universe with accurate prices. So create sheets first,
    # write price history, then come back to universe.

    # Sheet 2: Price History
    ws_prices = wb.create_sheet("Price History")
    logger.info("Writing price history (this may take a moment)...")
    _write_price_history_sheet(ws_prices, tickers, profiles, rng)

    # Now profiles have updated ending prices - write universe
    for i, (tk, prof) in enumerate(zip(tickers, profiles)):
        row = i + 2  # row 1 is header
        dollar_vol = round(prof["avg_volume"] * prof["price"], 2)
        ws_universe.cell(row=row, column=1, value=tk)
        ws_universe.cell(row=row, column=2, value=round(prof["market_cap"], 1))
        ws_universe.cell(row=row, column=3, value=prof["exchange"])
        ws_universe.cell(row=row, column=4, value=round(prof["avg_volume"]))
        ws_universe.cell(row=row, column=5, value=prof["price"])
        ws_universe.cell(row=row, column=6, value=dollar_vol)

    # Sheet 3: Fundamentals
    ws_fund = wb.create_sheet("Fundamentals")
    logger.info("Writing fundamentals...")
    _write_fundamentals_sheet(ws_fund, tickers, profiles, rng)

    # Sheet 4: Scores (empty - computed by engine)
    wb.create_sheet("Scores")

    # Sheet 5: Portfolio (empty - computed by engine)
    wb.create_sheet("Portfolio")

    # --- Save ---
    logger.info("Saving workbook to %s", output_path)
    wb.save(str(output_path))
    logger.info("Mock workbook saved: %d tickers, %d sheets", n_tickers, len(wb.sheetnames))

    return output_path


# ---------------------------------------------------------------------------
# CLI convenience
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    out = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/mock_bloomberg.xlsx"
    path = generate_mock_data(out)
    print(f"Created: {path}")
